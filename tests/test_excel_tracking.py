from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from po_label_app.excel_io import load_input_workbook
from po_label_app.models import ProcessingResult, STATUS_REQUESTED
from po_label_app.tracking import TrackingWorkbook


class ExcelTrackingTests(unittest.TestCase):
    def test_input_validation_trims_and_deduplicates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "input.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["PO", "Printer"])
            sheet.append([" 00123 ", " hk "])
            sheet.append(["", "VN"])
            sheet.append(["00123", "HK"])
            workbook.save(path)

            loaded = load_input_workbook(path, "run001")

            self.assertEqual(len(loaded.orders), 1)
            self.assertEqual(loaded.orders[0].po, "00123")
            self.assertEqual(loaded.orders[0].printer, "hk")
            self.assertEqual(len(loaded.duplicate_results), 1)

    def test_tracking_persists_requested_pos(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "tracking.xlsx"
            tracking = TrackingWorkbook(path)
            tracking.append_result(
                ProcessingResult.create(
                    po="5946692",
                    printer="HK",
                    automation_status=STATUS_REQUESTED,
                    error_explanation="",
                    run_id="run001",
                )
            )

            reopened = TrackingWorkbook(path)

            self.assertEqual(reopened.requested_pos(), {"5946692"})


if __name__ == "__main__":
    unittest.main()
