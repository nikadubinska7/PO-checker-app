from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .models import ProcessingResult, PurchaseOrder, STATUS_SKIPPED


REQUIRED_COLUMNS = {"PO", "Printer"}
TRACKING_COLUMNS = [
    "PO",
    "Printer",
    "Automation Status",
    "Error Explanation",
    "Processed Date",
    "Processed Timestamp",
    "Screenshot Path",
    "Run ID",
]


@dataclass(frozen=True)
class InputWorkbook:
    orders: list[PurchaseOrder]
    duplicate_results: list[ProcessingResult]


def _normalize_header(value: object) -> str:
    return str(value or "").strip()


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_input_workbook(path: str | Path, run_id: str) -> InputWorkbook:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Input Excel file does not exist: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [_normalize_header(cell.value) for cell in sheet[1]]
    header_map = {header: index for index, header in enumerate(headers) if header}

    missing = REQUIRED_COLUMNS - set(header_map)
    if missing:
        missing_text = ", ".join(sorted(missing))
        raise ValueError(f"Input Excel file is missing required column(s): {missing_text}")

    orders: list[PurchaseOrder] = []
    duplicate_results: list[ProcessingResult] = []
    seen: set[str] = set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        po = _normalize_cell(row[header_map["PO"]])
        printer = _normalize_cell(row[header_map["Printer"]])
        if not po:
            continue
        if po in seen:
            duplicate_results.append(
                ProcessingResult.create(
                    po=po,
                    printer=printer,
                    automation_status=STATUS_SKIPPED,
                    error_explanation="Duplicate PO in input run; first occurrence processed",
                    run_id=run_id,
                )
            )
            continue
        seen.add(po)
        orders.append(PurchaseOrder(po=po, printer=printer, source_row=row_number))

    return InputWorkbook(orders=orders, duplicate_results=duplicate_results)
