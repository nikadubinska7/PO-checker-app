from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .excel_io import TRACKING_COLUMNS
from .models import ProcessingResult, STATUS_REQUESTED


class TrackingWorkbook:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            self.workbook = load_workbook(self.path)
            self.sheet = self.workbook.active
            self._ensure_headers()
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Tracking"
            self.sheet.append(TRACKING_COLUMNS)
            self.save()

    def requested_pos(self) -> set[str]:
        column_map = self._column_map()
        po_index = column_map.get("PO")
        status_index = column_map.get("Automation Status")
        if po_index is None or status_index is None:
            return set()
        requested: set[str] = set()
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            po = str(row[po_index] or "").strip()
            status = str(row[status_index] or "").strip().lower()
            if po and status == STATUS_REQUESTED:
                requested.add(po)
        return requested

    def append_result(self, result: ProcessingResult) -> None:
        self.sheet.append(
            [
                result.po,
                result.printer,
                result.automation_status,
                result.error_explanation,
                result.processed_date,
                result.processed_timestamp,
                result.screenshot_path,
                result.run_id,
            ]
        )
        self._format_sheet()
        self.save()

    def append_results(self, results: list[ProcessingResult]) -> None:
        for result in results:
            self.append_result(result)

    def save(self) -> None:
        self.workbook.save(self.path)

    def _column_map(self) -> dict[str, int]:
        headers = [str(cell.value or "").strip() for cell in self.sheet[1]]
        return {header: index for index, header in enumerate(headers) if header}

    def _ensure_headers(self) -> None:
        if self.sheet.max_row == 0:
            self.sheet.append(TRACKING_COLUMNS)
            return
        existing = [str(cell.value or "").strip() for cell in self.sheet[1]]
        if not any(existing):
            for index, header in enumerate(TRACKING_COLUMNS, start=1):
                self.sheet.cell(row=1, column=index, value=header)
            return
        for header in TRACKING_COLUMNS:
            if header not in existing:
                self.sheet.cell(row=1, column=len(existing) + 1, value=header)
                existing.append(header)

    def _format_sheet(self) -> None:
        self.sheet.freeze_panes = "A2"
        widths = {
            "A": 16,
            "B": 12,
            "C": 20,
            "D": 42,
            "E": 16,
            "F": 22,
            "G": 60,
            "H": 28,
        }
        for column, width in widths.items():
            self.sheet.column_dimensions[column].width = width


def ensure_tracking_path(path: str | Path) -> Path:
    tracking_path = Path(path)
    if tracking_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Tracking output file must be an Excel .xlsx or .xlsm file")
    return tracking_path
